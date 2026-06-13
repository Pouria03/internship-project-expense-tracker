import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from apps.expenses.helpers.jalali_helper import to_jalali


def export_expenses_to_excel(expenses):
    wb = Workbook()
    ws = wb.active
    ws.title = 'گزارش هزینه‌ها'
    ws.sheet_view.rightToLeft = True

    # Header
    headers = ['#', 'عنوان', 'دسته‌بندی', 'مقدار', 'مبلغ (تومان)', 'کل (تومان)', 'ارزیابی', 'توضیحات', 'تاریخ']
    header_fill = PatternFill(start_color='36A2EB', end_color='36A2EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=expense.title)
        ws.cell(row=row, column=3, value=expense.category.name)
        ws.cell(row=row, column=4, value=expense.quantity)
        ws.cell(row=row, column=5, value=int(expense.amount))
        ws.cell(row=row, column=6, value=int(expense.total))
        ws.cell(row=row, column=7, value=expense.get_evaluation_display())
        ws.cell(row=row, column=8, value=expense.description)
        ws.cell(row=row, column=9, value=to_jalali(expense.created_at))

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 15

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output